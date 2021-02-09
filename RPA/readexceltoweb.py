from openpyxl import load_workbook
from selenium import webdriver
import time

excelfile = load_workbook('fruit.xlsx')
sheet = excelfile.active # file current

print(sheet['B1'].value) # read cell B1

alltitle = []
allprice = []
alldata = []

for i in range(3,7):
	pdname = sheet.cell(row=i, column=2).value # read cell
	pdprice = sheet.cell(row=i, column=3).value
	pddetail = sheet.cell(row=i, column=4).value
	alltitle.append(pdname)
	allprice.append(pdprice)
	alldata.append(pddetail)

login='http://www.uncle.com/login/'
addpage = 'http://www.uncle.com/addproduct/'

web=webdriver.Chrome()
web.get(login)

time.sleep(1)

email=web.find_element_by_id('username')
email.send_keys('@gmail.com')
time.sleep(1)
pw=web.find_element_by_id('password')
pw.send_keys('1234')

'''
to use selenium package to press enter:
from selenium.webdriver.common.keys import Keys
pw.send_keys(Keys.RETURN)
'''

bt=web.find_element_by_xpath('/html/body/div[2]/form/button')
bt.click()
time.sleep(1)

# Add products
web.get(addpage)
time.sleep(2)

for n,p,d in zip(alltitle,allprice,alldata):

	addname=web.find_element_by_id('name')
	addname.send_keys(n)
	time.sleep(1)
	addprice=web.find_element_by_id('price')
	addprice.send_keys(p)
	time.sleep(1)
	adddetail=web.find_element_by_id('detail')
	adddetail.send_keys(d)
	time.sleep(3)
	bt=web.find_element_by_xpath('/html/body/div[2]/form/button')
	bt.click()
	time.sleep(1)
